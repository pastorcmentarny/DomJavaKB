import datetime
import random

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from openpyxl.utils import get_column_letter

'''
First, let’s go over some basic definitions: 
An Excel spreadsheet document is called a workbook. 
A single workbook is saved in a file with the .xlsx extension.
Each workbook can contain multiple sheets (also called worksheets). 
The sheet the user is currently viewing (or last viewed before closing Excel) is called the active sheet.

Each sheet has columns (addressed by letters starting at A) and rows (addressed by numbers starting at 1).
A box at a particular column and row is called a cell.
Each cell can contain a number or text value.
The grid of cells with data makes up a sheet.
'''


def displayInfoFor(cell):
    print('Cell [' + str(cell.row) + ':' + str(cell.column) + '] : Value: ' + str(cell.value))


base = 'C:\\ds\\projects\\DsPythonKB\\resources\\'
workbook = openpyxl.load_workbook(base + 'example.xlsx')
print(type(workbook))

for sheet in workbook.get_sheet_names():
    print('Sheet: ' + sheet)

sheet = workbook.get_sheet_by_name('Sheet3')
print(type(sheet))
activeSheet = workbook.active
print(activeSheet)

sheet = workbook.get_sheet_by_name('Sheet1')
print('Type of sheet[A1]: ')
print(type(sheet['A1']))

print(sheet['A1'].value)
cell = sheet['B1']
print(cell.value)
print("Cell info: Row:" + str(cell.row) + ' Column: ' + str(cell.column) + ' Value:' + cell.value)
displayInfoFor(cell)
displayInfoFor(sheet['A3'])
displayInfoFor(sheet.cell(row=1, column=2))
print(str(sheet.cell(row=2, column=2).value))

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(get_column_letter(1))
print(column_index_from_string('AA'))

sheet = workbook.get_sheet_by_name('Sheet1')
print(tuple(sheet['A1':'C3']))

for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

workbook = openpyxl.Workbook()
names = workbook.get_sheet_names()
print(names)

sheet = workbook.active
print(sheet.title)

sheet.title = 'I am hungry'
now = datetime.datetime.now()
date_separator = '-'
today = str(now.day) + date_separator + str(now.month) + date_separator + str(now.year)
print(today)
print(workbook.get_sheet_names())
workbook.save(base + 'copy_of_example_' + today + '.xlsx')
print(workbook.get_sheet_names());

# fix it
# sheet = workbook.active
# print(sheet.columns(1))
# for cellObj in sheet.columns[1]:
# print(cellObj.value)

workbook = openpyxl.Workbook()
workSheetTitle = 'Middle Sheet'

print("Creating and Removing Sheets")
print(workbook.get_sheet_names())
workbook.create_chartsheet('Food is good')
print(workbook.get_sheet_names())
workbook.create_chartsheet(index=0, title='First sheet')
print(workbook.get_sheet_names())
workbook.create_chartsheet(index=2, title='Middle Sheet')
print(workbook.get_sheet_names())
workbook.save(base + 'copy_of_example_' + today + 'part2.xlsx')
'''
DOES NOT WORK
shit = workbook.get_sheet_by_name('Food is good') 
print(type(shit))
workbook.remove_sheet(shit)
print(workbook.get_sheet_names())
'''

workbook = openpyxl.Workbook()
sheet = workbook.get_sheet_by_name('Sheet')
sheet['E8'] = 'UFO'
print(sheet['E8'].value)
workbook.save(base + 'copy_of_example_' + today + 'part3.xlsx')

workbook.close()

'''
The row height can be set to an integer or float value between 0 and 409.
This value represents the height measured in points, where one point equals 1/72 of an inch.
The default row height is 12.75. 
The column width can be set to an integer or float value between 0 and 255. 
'''

workbook = openpyxl.Workbook()
sheet = workbook.get_sheet_by_name('Sheet')
titleStyle = Font(size=36, italic=True, bold=True)
whatStyle = Font(size=12, bold=True)
priceStyle = Font(size=12, italic=True)
totalPayStyle = Font(size=18, italic=True, bold=True)

sheet = workbook.active
sheet['A1'].font = titleStyle
sheet['A1'] = 'Daily spent'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['A'].width = 40
sheet.freeze_panes = "A2"  # froze row 1 , 'B1' froze Column A , Row 1 and columns A and B, sheet.freeze_panes = None
# Frozen column or row headers, for example, are always visible to the user even as they scroll through the spreadsheet.
sheet['A2'].font = titleStyle
sheet['A2'] = "Commute"

sheet['B2'].font = priceStyle
sheet['B2'] = 11.2

sheet['A3'].font = titleStyle
sheet['A3'] = "Coffee"

sheet['B3'].font = priceStyle
sheet['B3'] = 2.7

sheet['A4'].font = titleStyle
sheet['A4'] = "Other"

sheet['B4'].font = priceStyle
sheet['B4'] = 5.6

sheet['A5'] = "-- --- ----- -------- "
sheet.merge_cells('A5:D5')

sheet['A6'].font = totalPayStyle
sheet['A6'] = "Total:"

sheet['B6'].font = totalPayStyle
sheet['B6'] = '=SUM(B2:B4)'

sheet['A7'] = "Please spend your money responsibly"

mergeToUnmerge = 'A9:D9'
sheet.merge_cells(mergeToUnmerge)
sheet.unmerge_cells(mergeToUnmerge)

workbook.save(base + 'styled_example_' + today + 'part' + str(random.randint(0, 100)) + '.xlsx')

# you can create styles and then use them
